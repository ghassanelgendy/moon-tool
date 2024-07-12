import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime  # Added for filename timestamp



timestamp = datetime.now().strftime("%d_%H_%M")  # Get current timestamp
output_filename = f"CSAT {timestamp}.xlsx"
def process_and_export_to_excel(file_path, output_path):


    # Load the CSV file
    data = pd.read_csv(file_path)

    # Remove leading/trailing spaces from column names
    data.columns = data.columns.str.strip()

    # Replace 'No_Answer' with 0 and convert 'Answer' to numeric
    data['Answer'] = data['Answer'].replace('No_Answer', 0)
    data['Answer'] = pd.to_numeric(data['Answer'], errors='coerce')

    # Drop rows where 'Answer' is NaN
    data = data.dropna(subset=['Answer'])

    # Ensure 'Answer' column is integer
    data['Answer'] = data['Answer'].astype(int)

    # Create the pivot table
    pivot_table = data.pivot_table(index='Agent Name', columns='Answer', values='Customer Phone Number', aggfunc='count', fill_value=0).reset_index()

    # Rename columns
    pivot_table.columns.name = None
    pivot_table = pivot_table.rename(columns={0: 'No_Answer', 1: 'Good', 2: 'Bad'})
    
    # Drop the 'No_Answer' column
    pivot_table = pivot_table.drop(columns=['No_Answer'])
    
    # Create the Surveys column
    pivot_table['Surveys'] = pivot_table['Good'] + pivot_table['Bad']
    
    # Create the CSAT column
    pivot_table['CSAT'] = (pivot_table['Good'] / pivot_table['Surveys']).apply(lambda x: f"{x:.0%}" if not pd.isna(x) else None)
    
    # Drop rows with NaN CSAT
    pivot_table = pivot_table.dropna(subset=['CSAT'])
    
    # Convert CSAT to numeric for sorting and conditional formatting
    pivot_table['CSAT_numeric'] = pivot_table['CSAT'].str.rstrip('%').astype(float)
    
    # Sort by CSAT
    pivot_table = pivot_table.sort_values(by='CSAT_numeric', ascending=False)
    
    # Append the Grand Total row
    grand_total = pivot_table[['Good', 'Bad', 'Surveys']].sum()
    grand_total['Agent Name'] = 'Grand Total'
    grand_total['CSAT'] = f"{grand_total['Good'] / grand_total['Surveys']:.0%}"
    grand_total['CSAT_numeric'] = float(grand_total['CSAT'].rstrip('%'))
    grand_total = pd.DataFrame(grand_total).transpose()
    pivot_table = pd.concat([pivot_table, grand_total], ignore_index=True)
    
    # Export to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pivot Table'
    
    # Write DataFrame to Excel
    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    
    # Center align all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


    # Adjust the width of the 'Agent Name' column to fit the longest name
    max_length = max(len(str(cell.value)) for cell in ws['A'] if cell.value)
    ws.column_dimensions['A'].width = max_length + 2  # Adding a little extra space for padding

    # Apply gradient color scale to the CSAT_numeric column
    first_data_row = 2
    last_data_row = ws.max_row - 1  # Exclude the Grand Total row from conditional formatting

    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        end_type='num', end_value=100, end_color='00FF00'
    )
    ws.conditional_formatting.add(f"F{first_data_row}:F{last_data_row}", rule)
    
    # Hide the CSAT_numeric column
    ws.column_dimensions['F'].hidden = True
    
    # Save the workbook
    wb.save(output_path)


# Usage to open a file starting with "IVR" in the same directory
import glob
ivr_file = glob.glob('IVR*.csv')[0]  # Assuming there's only one file starting with "IVR" in the directory
process_and_export_to_excel(ivr_file, output_filename)

#------------------------------========================================================
#------------------------------========================================================
#------------------------------========================================================
#------------------------------========================================================
