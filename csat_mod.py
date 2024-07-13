import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows

def process_and_export_to_excel(file_path, output_path):
    # Load the CSV file
    data = pd.read_csv(file_path)

    # Remove leading/trailing spaces from column names
    data.columns = data.columns.str.strip()

    # Replace 'No_Answer' with 0 and convert 'Answer' to numeric
    data['Answer'] = data['Answer'].replace('No_Answer', 0)
    data['Answer'] = pd.to_numeric(data['Answer'], errors='coerce')

    # Drop rows where 'Answer' is NaN
    data = data.dropna(subset=['Answer'])

    # Ensure 'Answer' column is integer
    data['Answer'] = data['Answer'].astype(int)

    # Create the pivot table
    pivot_table = data.pivot_table(index='Agent Name', columns='Answer', values='Customer Phone Number', aggfunc='count', fill_value=0).reset_index()

    # Rename columns
    pivot_table.columns.name = None
    pivot_table = pivot_table.rename(columns={0: 'No_Answer', 1: 'Good', 2: 'Bad'})
    
    # Drop the 'No_Answer' column
    pivot_table = pivot_table.drop(columns=['No_Answer'])
    
    # Create the Surveys column
    pivot_table['Surveys'] = pivot_table['Good'] + pivot_table['Bad']
    
    # Create the CSAT column as a percentage
    pivot_table['CSAT'] = (pivot_table['Good'] / pivot_table['Surveys']) * 100
    
    # Drop rows with NaN CSAT
    pivot_table = pivot_table.dropna(subset=['CSAT'])
    
    # Sort by CSAT
    pivot_table = pivot_table.sort_values(by='CSAT', ascending=False)
    
    # Append the Grand Total row
    grand_total = pivot_table[['Good', 'Bad', 'Surveys']].sum()
    grand_total['Agent Name'] = 'Grand Total'
    grand_total['CSAT'] = (grand_total['Good'] / grand_total['Surveys']) * 100
    grand_total = pd.DataFrame(grand_total).transpose()
    pivot_table = pd.concat([pivot_table, grand_total], ignore_index=True)
    
    # Export to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pivot Table'
    
    # Write DataFrame to Excel
    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='5FACC7', end_color='5FACC7', fill_type='solid')
    
    # Center align all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Adjust the width of the 'Agent Name' column to fit the longest name
    max_length = max(len(str(cell.value)) for cell in ws['A'] if cell.value)
    ws.column_dimensions['A'].width = max_length + 2  # Adding a little extra space for padding

    # Apply gradient color scale to the CSAT column
    first_data_row = 2
    last_data_row = ws.max_row - 1  # Exclude the Grand Total row from conditional formatting

    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',  # Red for the lowest value
        mid_type='num', mid_value=50, mid_color='FFFF00',       # Yellow for the middle value
        end_type='num', end_value=100, end_color='00FF00'       # Green for the highest value
    )
    ws.conditional_formatting.add(f"E{first_data_row}:E{last_data_row}", rule)
    
    # Format the CSAT column as integer percentage
    for cell in ws[f"E{first_data_row}:E{ws.max_row}"]:
        for c in cell:
            c.number_format = '0%'

    # Apply background color to the Grand Total row
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='5FACC7', end_color='5FACC7', fill_type='solid')
    
    # Save the workbook
    wb.save(output_path)

# Example usage
file_path = "hh.csv"
output_path = "two.xlsx"
process_and_export_to_excel(file_path, output_path)
