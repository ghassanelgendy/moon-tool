import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def read_csv_skip_rows(file_path, skip_rows=4):
    # Read the CSV file and skip the first 4 rows
    df = pd.read_csv(file_path, skiprows=skip_rows)
    return df

def filter_by_hour(df, hour):
    # Convert "Ticket Closed Time" to datetime with the correct format
    df['Ticket Closed Time'] = pd.to_datetime(df['Ticket Closed Time'], format='%d %b %Y %I:%M %p')
    
    # Filter the DataFrame to only include rows where "Ticket Closed Time" is within the specified hour
    filtered_df = df[df['Ticket Closed Time'].dt.hour == hour]
    return filtered_df

def create_pivot_table(df):
    # Create a pivot table similar to the provided image
    pivot_table = pd.pivot_table(df, index='Ticket Owner', columns='Team', values='Ticket Id', aggfunc='count', margins=True, margins_name='Grand Total', fill_value=0)
    
    # Sort the pivot table rows by the row totals in descending order (excluding 'Grand Total' row)
    pivot_table = pivot_table.sort_values(by='Grand Total', ascending=False)
    
    # Move 'Grand Total' row to the last row
    grand_total_row = pivot_table.loc['Grand Total']
    pivot_table = pivot_table.drop(index='Grand Total')
    pivot_table = pd.concat([pivot_table, grand_total_row.to_frame().T])  # Concatenate the pivot table with grand_total_row

    return pivot_table

def style_pivot_table(ws):
    # Apply some basic styling to the pivot table
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1 or cell.column == 1 or cell.row == ws.max_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            cell.border = thin_border

def save_to_excel(df, pivot_table, output_file):
    # Create a new workbook
    wb = Workbook()

    # Write the filtered data to the first sheet
    ws_filtered = wb.active
    ws_filtered.title = "Filtered Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_filtered.append(r)

    # Create a new sheet for the pivot table
    ws_pivot = wb.create_sheet(title="Pivot Table")

    # Write the pivot table to the new sheet
    pivot_table.reset_index(inplace=True)
    pivot_table.columns.name = None  # Remove the name of the columns
    pivot_table.rename(columns={'index': 'Ticket Owner'}, inplace=True)

    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws_pivot.append(r)

    # Apply styling to the pivot table
    style_pivot_table(ws_pivot)

    # Adjust column widths
    for col in ws_pivot.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_pivot.column_dimensions[column].width = adjusted_width

    # Reorder sheets so that Pivot Table sheet is the first one
    wb.move_sheet(ws_filtered, offset=1)  # Move Filtered Data sheet to the second position

    # Save the workbook to a file
    wb.save(output_file)

def automate_process(csv_file, hour, output_file):
    df = read_csv_skip_rows(csv_file)
    filtered_df = filter_by_hour(df, hour)
    pivot_table = create_pivot_table(filtered_df)
    save_to_excel(filtered_df, pivot_table, output_file)

def main():
    print("Make sure the file name is 'ghassan' :)")
    csv_file = 'L2 UAE Intraday.csv'
    hour =  11
    output_file = 'Prod elsa3a '+ str(hour) +' yabasha.xlsx'
    automate_process(csv_file, hour, output_file)

if __name__ == '__main__':
    main()
