@ -1,354 +0,0 @@
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime  # Added for filename timestamp
import glob
import os


timestamp = datetime.now().strftime("%d_%H_%M")  # Get current timestamp
output_filename = f"CSAT {timestamp}.xlsx"
def process_and_export_to_excel(file_path, output_path):


    # Load the CSV file
    data = pd.read_csv(file_path)

    # Remove leading/trailing spaces from column names
    data.columns = data.columns.str.strip()

    # Replace 'No_Answer' with 0 and convert 'Answer' to numeric
    data['Answer'] = data['Answer'].replace('No_Answer', 0)
    data['Answer'] = pd.to_numeric(data['Answer'], errors='coerce')

    # Drop rows where 'Answer' is NaN
    data = data.dropna(subset=['Answer'])

    # Ensure 'Answer' column is integer
    data['Answer'] = data['Answer'].astype(int)

    # Create the pivot table
    pivot_table = data.pivot_table(index='Agent Name', columns='Answer', values='Customer Phone Number', aggfunc='count', fill_value=0).reset_index()

    # Rename columns
    pivot_table.columns.name = None
    pivot_table = pivot_table.rename(columns={0: 'No_Answer', 1: 'Good', 2: 'Bad'})
    
    # Drop the 'No_Answer' column
    pivot_table = pivot_table.drop(columns=['No_Answer'])
    
    # Create the Surveys column
    pivot_table['Surveys'] = pivot_table['Good'] + pivot_table['Bad']
    
    # Create the CSAT column
    pivot_table['CSAT'] = (pivot_table['Good'] / pivot_table['Surveys']).apply(lambda x: f"{x:.0%}" if not pd.isna(x) else None)
    
    # Drop rows with NaN CSAT
    pivot_table = pivot_table.dropna(subset=['CSAT'])
    
    # Convert CSAT to numeric for sorting and conditional formatting
    pivot_table['CSAT_numeric'] = pivot_table['CSAT'].str.rstrip('%').astype(float)
    
    # Sort by CSAT
    pivot_table = pivot_table.sort_values(by='CSAT_numeric', ascending=False)
    
    # Append the Grand Total row
    grand_total = pivot_table[['Good', 'Bad', 'Surveys']].sum()
    grand_total['Agent Name'] = 'Grand Total'
    grand_total['CSAT'] = f"{grand_total['Good'] / grand_total['Surveys']:.0%}"
    grand_total['CSAT_numeric'] = float(grand_total['CSAT'].rstrip('%'))
    grand_total = pd.DataFrame(grand_total).transpose()
    pivot_table = pd.concat([pivot_table, grand_total], ignore_index=True)
    
    # Export to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pivot Table'
    
    # Write DataFrame to Excel
    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    
    # Center align all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


    # Adjust the width of the 'Agent Name' column to fit the longest name
    max_length = max(len(str(cell.value)) for cell in ws['A'] if cell.value)
    ws.column_dimensions['A'].width = max_length + 2  # Adding a little extra space for padding

    # Apply gradient color scale to the CSAT_numeric column
    first_data_row = 2
    last_data_row = ws.max_row - 1  # Exclude the Grand Total row from conditional formatting

    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        end_type='num', end_value=100, end_color='00FF00'
    )
    ws.conditional_formatting.add(f"F{first_data_row}:F{last_data_row}", rule)
    
    # Hide the CSAT_numeric column
    ws.column_dimensions['F'].hidden = True
    
    # Save the workbook
    wb.save(output_path)

def read_csv_skip_rows(file_path, skip_rows=4):
    # Read the CSV file and skip the first 4 rows
    df = pd.read_csv(file_path, skiprows=skip_rows)
    return df

def filter_by_hour(df, hour):
    # Convert "Ticket Closed Time" to datetime with the correct format
    df['Ticket Closed Time'] = pd.to_datetime(df['Ticket Closed Time'], format='%d %b %Y %I:%M %p')
    
    # Filter the DataFrame to only include rows where "Ticket Closed Time" is within the specified hour
    filtered_df = df[df['Ticket Closed Time'].dt.hour == hour]
    return filtered_df

def filter_by_day(df, day):
    # Convert "Ticket Closed Time" to datetime with the correct format
    df['Ticket Closed Time'] = pd.to_datetime(df['Ticket Closed Time'], format='%d %b %Y %I:%M %p')

    # Filter the DataFrame to only include rows where "Ticket Closed Time" is within the specified hour
    filtered_df = df[df['Ticket Closed Time'].dt.day == day]
    return filtered_df
def create_pivot_table(df):
    # Create a pivot table similar to the provided image
    pivot_table = pd.pivot_table(df, index='Ticket Owner', columns='Team', values='Ticket Id', aggfunc='count', margins=True, margins_name='Grand Total', fill_value=0)
    
    # Sort the pivot table rows by the row totals in descending order (excluding 'Grand Total' row)
    pivot_table = pivot_table.sort_values(by='Grand Total', ascending=False)
    
    # Move 'Grand Total' row to the last row
    grand_total_row = pivot_table.loc['Grand Total']
    pivot_table = pivot_table.drop(index='Grand Total')
    pivot_table = pd.concat([pivot_table, grand_total_row.to_frame().T])  # Concatenate the pivot table with grand_total_row

    return pivot_table

def style_pivot_table(ws):
    # Apply some basic styling to the pivot table
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1 or cell.column == 1 or cell.row == ws.max_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            cell.border = thin_border

def save_to_excel(df, pivot_table, output_file):
    # Create a new workbook
    wb = Workbook()

    # Write the filtered data to the first sheet
    ws_filtered = wb.active
    ws_filtered.title = "Filtered Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_filtered.append(r)

    # Create a new sheet for the pivot table
    ws_pivot = wb.create_sheet(title="Pivot Table")

    # Write the pivot table to the new sheet
    pivot_table.reset_index(inplace=True)
    pivot_table.columns.name = None  # Remove the name of the columns
    pivot_table.rename(columns={'index': 'Ticket Owner'}, inplace=True)

    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws_pivot.append(r)

    # Apply styling to the pivot table
    style_pivot_table(ws_pivot)

    # Adjust column widths
    for col in ws_pivot.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_pivot.column_dimensions[column].width = adjusted_width

    # Reorder sheets so that Pivot Table sheet is the first one
    wb.move_sheet(ws_filtered, offset=1)  # Move Filtered Data sheet to the second position

    # Save the workbook to a file
    wb.save(output_file)

def automate_process(csv_file, hour, output_file):
    df = read_csv_skip_rows(csv_file)
    filtered_df = filter_by_hour(df, hour)
    pivot_table = create_pivot_table(filtered_df)
    save_to_excel(filtered_df, pivot_table, output_file)


timestamp = datetime.now().strftime("%d_%H_%M")  # Get current timestamp
output_filename = f"CSAT {timestamp}.xlsx"
def process_and_export_to_excel(file_path, output_path):


    # Load the CSV file
    data = pd.read_csv(file_path)

    # Remove leading/trailing spaces from column names
    data.columns = data.columns.str.strip()

    # Replace 'No_Answer' with 0 and convert 'Answer' to numeric
    data['Answer'] = data['Answer'].replace('No_Answer', 0)
    data['Answer'] = pd.to_numeric(data['Answer'], errors='coerce')

    # Drop rows where 'Answer' is NaN
    data = data.dropna(subset=['Answer'])

    # Ensure 'Answer' column is integer
    data['Answer'] = data['Answer'].astype(int)

    # Create the pivot table
    pivot_table = data.pivot_table(index='Agent Name', columns='Answer', values='Customer Phone Number', aggfunc='count', fill_value=0).reset_index()

    # Rename columns
    pivot_table.columns.name = None
    pivot_table = pivot_table.rename(columns={0: 'No_Answer', 1: 'Good', 2: 'Bad'})
    
    # Drop the 'No_Answer' column
    pivot_table = pivot_table.drop(columns=['No_Answer'])
    
    # Create the Surveys column
    pivot_table['Surveys'] = pivot_table['Good'] + pivot_table['Bad']
    
    # Create the CSAT column
    pivot_table['CSAT'] = (pivot_table['Good'] / pivot_table['Surveys']).apply(lambda x: f"{x:.0%}" if not pd.isna(x) else None)
    
    # Drop rows with NaN CSAT
    pivot_table = pivot_table.dropna(subset=['CSAT'])
    
    # Convert CSAT to numeric for sorting and conditional formatting
    pivot_table['CSAT_numeric'] = pivot_table['CSAT'].str.rstrip('%').astype(float)
    
    # Sort by CSAT
    pivot_table = pivot_table.sort_values(by='CSAT_numeric', ascending=False)
    
    # Append the Grand Total row
    grand_total = pivot_table[['Good', 'Bad', 'Surveys']].sum()
    grand_total['Agent Name'] = 'Grand Total'
    grand_total['CSAT'] = f"{grand_total['Good'] / grand_total['Surveys']:.0%}"
    grand_total['CSAT_numeric'] = float(grand_total['CSAT'].rstrip('%'))
    grand_total = pd.DataFrame(grand_total).transpose()
    pivot_table = pd.concat([pivot_table, grand_total], ignore_index=True)
    
    # Export to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pivot Table'
    
    # Write DataFrame to Excel
    for r in dataframe_to_rows(pivot_table, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    
    # Center align all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


    # Adjust the width of the 'Agent Name' column to fit the longest name
    max_length = max(len(str(cell.value)) for cell in ws['A'] if cell.value)
    ws.column_dimensions['A'].width = max_length + 2  # Adding a little extra space for padding

    # Apply gradient color scale to the CSAT_numeric column
    first_data_row = 2
    last_data_row = ws.max_row - 1  # Exclude the Grand Total row from conditional formatting

    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        end_type='num', end_value=100, end_color='00FF00'
    )
    ws.conditional_formatting.add(f"F{first_data_row}:F{last_data_row}", rule)
    
    # Hide the CSAT_numeric column
    ws.column_dimensions['F'].hidden = True
    
    # Save the workbook
    wb.save(output_path)


def main():
    while True:
        print("===========[[ Welcome to Ghassan's tool for cool TCs and RTMs ]]===========")
        print(
            ''' 
1) Productivity for an hour
2) Productivity for a day
3) C-SAT overday
4) Get help
5) Exit

More tools to be announced soon lw mamshetsh
            ''')
        choice = input("Enter the number of the tool you want to use: ")
        #choice = '2';
        if choice == '1':
            print("Make sure the file name is 'ghassan' :)")
            csv_file = 'ghassan.csv'
            hour = int(input("Enter the hour you want to filter by (0-23): "))
            output_file = 'Prod youm elsa3a ' + str(hour) + ' yabasha.xlsx'
            automate_process(csv_file, hour, output_file)
            print("Done! The output file has been saved as:", output_file)

        elif choice == '2':
            csv_file = glob.glob('L2*.csv')[0]
            day = int(input("Enter the day you want to filter by (1-31): "))
            output_file = 'Prod elyom ' + str(day) + ' yabasha.xlsx'
            df = read_csv_skip_rows(csv_file)
            filtered_df = filter_by_day(df, day)
            pivot_table = create_pivot_table(filtered_df)
            save_to_excel(filtered_df, pivot_table, output_file)
            print("Done! The output file has been saved as:", output_file)

        elif choice == '3':
# Usage to open a file starting with "IVR" in the same directory
            ivr_file = glob.glob('IVR*.csv')[0]  # Assuming there's only one file starting with "IVR" in the directory
            process_and_export_to_excel(ivr_file, output_filename)
            print("Done! The output file has been saved as:", output_filename)

        elif choice == '4':
            print("For productivity, please make sure the file name is 'ghassan' and the hour is set to the hour you want to start from,"
                  "\tfor example:"
                  "\thour = 16 for productivity from 4PM - 5PM."
                  "\thour = 3 for productivity from 3AM - 4AM.")
            print("For C-SAT, please make sure the file name is left as it is (starts with IVR) and is in the same folder as this file."
                  "\tThe tool will automatically filter the data for the specified hour and generate the output file with name : CSAT date_time")
            print("If you need further assistance, please contact Ghassan.")
        elif choice == '5':
            print("Exiting the tool. Salam")
            break
        else:
            print("Invalid choice. Please enter a valid number.")
        
        os.system("PAUSE")

if __name__ == '__main__':
    main()